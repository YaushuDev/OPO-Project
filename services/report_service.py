# report_service.py
"""
Servicio para generar reportes Excel optimizados de perfiles de búsqueda.
Crea archivos Excel con información esencial, seguimiento de ejecuciones óptimas
con formato condicional por rangos y generación de reportes diarios, semanales
y mensuales con cálculo correcto de porcentajes de éxito basados en objetivos.
"""

import os
import glob
from pathlib import Path
from datetime import datetime, timedelta, date
import calendar

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    openpyxl = None


class ReportService:
    """Servicio para generar reportes optimizados en formato Excel con cálculo corregido de éxito semanal y mensual."""

    def __init__(self):
        """Inicializa el servicio de reportes."""
        self.reports_dir = Path("reports")
        os.makedirs(self.reports_dir, exist_ok=True)

    def generate_profiles_report(self, profiles):
        """
        Genera un reporte Excel diario con información esencial de perfiles.

        Args:
            profiles (list): Lista de perfiles de búsqueda

        Returns:
            str: Ruta del archivo Excel generado
        """
        if openpyxl is None:
            raise Exception("openpyxl no está instalado. Ejecute: pip install openpyxl")

        # Crear nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_perfiles_{timestamp}.xlsx"
        file_path = self.reports_dir / filename

        # Crear libro de trabajo
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Perfiles de Búsqueda"

        # Configurar estilos
        styles = self._get_report_styles()

        # Crear título general
        self._add_daily_header(worksheet, len(profiles), styles)

        # Configurar encabezados de tabla
        self._add_table_headers(worksheet, styles)

        # Escribir datos de perfiles
        self._add_profile_data(worksheet, profiles, styles)

        # Ajustar formato
        self._format_daily_worksheet(worksheet)

        # Agregar hoja de resumen
        summary_sheet = workbook.create_sheet("Resumen")
        self._add_summary_sheet(summary_sheet, profiles)

        # Guardar archivo
        workbook.save(file_path)
        return str(file_path)

    def generate_weekly_profiles_report(self):
        """
        Genera un reporte semanal con cálculo corregido de porcentajes de éxito.
        Multiplica las ejecuciones óptimas diarias por 7 para obtener el objetivo semanal.

        Returns:
            str: Ruta del archivo Excel generado
        """
        if openpyxl is None:
            raise Exception("openpyxl no está instalado. Ejecute: pip install openpyxl")

        # Obtener fechas de la semana actual
        today = datetime.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        # Buscar y procesar reportes de la semana
        weekly_data = self._process_weekly_reports(start_of_week, end_of_week)

        if not weekly_data['reports_found']:
            raise Exception("No se encontraron reportes diarios para la semana actual")

        # Crear archivo de reporte semanal
        file_path = self._create_weekly_file(start_of_week, end_of_week)

        # Generar reporte
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reporte Semanal"

        # Configurar estilos
        styles = self._get_report_styles()

        # Crear contenido del reporte
        self._add_weekly_header(worksheet, start_of_week, end_of_week, len(weekly_data['aggregated_data']), styles)
        self._add_table_headers(worksheet, styles)
        self._add_weekly_profile_data(worksheet, weekly_data['aggregated_data'], styles)
        self._format_daily_worksheet(worksheet)

        # Agregar resumen semanal
        summary_sheet = workbook.create_sheet("Resumen Semanal")
        self._add_weekly_summary_sheet(summary_sheet, weekly_data, start_of_week, end_of_week)

        workbook.save(file_path)
        return str(file_path)

    def generate_monthly_profiles_report(self):
        """
        Genera un reporte mensual con cálculo corregido de porcentajes de éxito.
        Multiplica las ejecuciones óptimas diarias por los días del mes para obtener el objetivo mensual.

        Returns:
            str: Ruta del archivo Excel generado
        """
        if openpyxl is None:
            raise Exception("openpyxl no está instalado. Ejecute: pip install openpyxl")

        # Obtener fechas del mes actual
        today = datetime.now().date()
        # Primer día del mes
        start_of_month = date(today.year, today.month, 1)
        # Último día del mes
        _, last_day = calendar.monthrange(today.year, today.month)
        end_of_month = date(today.year, today.month, last_day)

        # Buscar y procesar reportes del mes
        monthly_data = self._process_monthly_reports(start_of_month, end_of_month)

        if not monthly_data['reports_found']:
            raise Exception("No se encontraron reportes diarios para el mes actual")

        # Crear archivo de reporte mensual
        file_path = self._create_monthly_file(start_of_month, end_of_month)

        # Generar reporte
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reporte Mensual"

        # Configurar estilos
        styles = self._get_report_styles()

        # Crear contenido del reporte
        self._add_monthly_header(worksheet, start_of_month, end_of_month, len(monthly_data['aggregated_data']), styles)
        self._add_table_headers(worksheet, styles)
        self._add_monthly_profile_data(worksheet, monthly_data['aggregated_data'], styles)
        self._format_daily_worksheet(worksheet)

        # Agregar resumen mensual
        summary_sheet = workbook.create_sheet("Resumen Mensual")
        self._add_monthly_summary_sheet(summary_sheet, monthly_data, start_of_month, end_of_month)

        workbook.save(file_path)
        return str(file_path)

    def _get_header_map(self, worksheet):
        """Obtiene un mapa de encabezados a columnas para manejo flexible de reportes."""
        header_map = {}
        header_row = 4
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=header_row, column=col).value
            if isinstance(value, str) and value.strip():
                header_map[value.strip().lower()] = col
        return header_map

    def _process_weekly_reports(self, start_of_week, end_of_week):
        """Procesa los reportes diarios de la semana y agrega los datos correctamente."""
        weekly_reports = []
        pattern = str(self.reports_dir / "reporte_perfiles_*.xlsx")

        # Buscar archivos de la semana
        for file_path in glob.glob(pattern):
            try:
                file_name = os.path.basename(file_path)
                date_part = file_name.split('_')[2].split('.')[0][:8]
                file_date = datetime.strptime(date_part, "%Y%m%d").date()

                if start_of_week <= file_date <= end_of_week:
                    weekly_reports.append(file_path)
            except (ValueError, IndexError):
                continue

        # Procesar datos agregados
        aggregated_data = {}

        for report_path in weekly_reports:
            try:
                wb = openpyxl.load_workbook(report_path, data_only=True)
                ws = wb.active

                header_map = self._get_header_map(ws)
                name_col = header_map.get("nombre del perfil")
                executions_col = header_map.get("cantidad de ejecuciones")
                optimal_col = header_map.get("cantidad de ejecuciones recomendadas")
                automatic_col = header_map.get("bot automático")
                manual_col = header_map.get("bot manual")
                offline_col = header_map.get("bot offline")
                last_search_col = header_map.get("última búsqueda")
                responsable_col = header_map.get("responsable")
                last_update_col = header_map.get("última actualización")
                delivery_col = header_map.get("fecha de entrega")

                if not name_col or not executions_col or not optimal_col:
                    continue

                for row in range(5, ws.max_row + 1):
                    profile_name = ws.cell(row=row, column=name_col).value
                    if not profile_name:
                        continue

                    responsable = ws.cell(row=row, column=responsable_col).value if responsable_col else ""
                    executions = ws.cell(row=row, column=executions_col).value or 0
                    optimal_cell = ws.cell(row=row, column=optimal_col).value
                    is_automatic = bool(automatic_col and ws.cell(row=row, column=automatic_col).value == "X")
                    is_manual = bool(manual_col and ws.cell(row=row, column=manual_col).value == "X")
                    is_offline = bool(offline_col and ws.cell(row=row, column=offline_col).value == "X")
                    last_search = ws.cell(row=row, column=last_search_col).value if last_search_col else None
                    last_update_text = ws.cell(row=row, column=last_update_col).value if last_update_col else ""
                    delivery_date_text = ws.cell(row=row, column=delivery_col).value if delivery_col else ""

                    if isinstance(last_update_text, str) and last_update_text.strip() == "—":
                        last_update_text = ""
                    if isinstance(delivery_date_text, str) and delivery_date_text.strip() == "—":
                        delivery_date_text = ""

                    # Procesar ejecuciones óptimas
                    daily_optimal = self._extract_optimal_value(optimal_cell)
                    has_tracking = daily_optimal > 0

                    if profile_name not in aggregated_data:
                        aggregated_data[profile_name] = {
                            'executions': 0,
                            'daily_optimal': daily_optimal,
                            'weekly_optimal': daily_optimal * 7 if has_tracking else 0,
                            'has_tracking': has_tracking,
                            'is_automatic': is_automatic,
                            'is_manual': is_manual,
                            'is_offline': is_offline,
                            'last_search': last_search,
                            'responsable': responsable,
                            'last_update_text': last_update_text or "",
                            'delivery_date_text': delivery_date_text or ""
                        }
                    else:
                        if responsable and not aggregated_data[profile_name].get('responsable'):
                            aggregated_data[profile_name]['responsable'] = responsable
                        aggregated_data[profile_name]['is_automatic'] = (
                            aggregated_data[profile_name]['is_automatic'] or is_automatic
                        )
                        aggregated_data[profile_name]['is_manual'] = (
                            aggregated_data[profile_name]['is_manual'] or is_manual
                        )
                        aggregated_data[profile_name]['is_offline'] = (
                            aggregated_data[profile_name].get('is_offline', False) or is_offline
                        )

                        if last_update_text:
                            aggregated_data[profile_name]['last_update_text'] = last_update_text

                        if delivery_date_text:
                            aggregated_data[profile_name]['delivery_date_text'] = delivery_date_text

                    # Acumular ejecuciones
                    aggregated_data[profile_name]['executions'] += executions

                    # Actualizar última búsqueda si es más reciente
                    if last_search and (not aggregated_data[profile_name]['last_search'] or
                                        last_search > aggregated_data[profile_name]['last_search']):
                        aggregated_data[profile_name]['last_search'] = last_search

            except Exception as e:
                print(f"Error procesando archivo {report_path}: {e}")
                continue

        return {
            'aggregated_data': aggregated_data,
            'reports_found': len(weekly_reports),
            'reports_count': len(weekly_reports)
        }

    def _process_monthly_reports(self, start_of_month, end_of_month):
        """Procesa los reportes diarios del mes y agrega los datos correctamente."""
        monthly_reports = []
        pattern = str(self.reports_dir / "reporte_perfiles_*.xlsx")

        # Buscar archivos del mes
        for file_path in glob.glob(pattern):
            try:
                file_name = os.path.basename(file_path)
                date_part = file_name.split('_')[2].split('.')[0][:8]
                file_date = datetime.strptime(date_part, "%Y%m%d").date()

                if start_of_month <= file_date <= end_of_month:
                    monthly_reports.append(file_path)
            except (ValueError, IndexError):
                continue

        # Procesar datos agregados
        aggregated_data = {}

        # Calcular el número de días en el mes
        days_in_month = (end_of_month - start_of_month).days + 1

        for report_path in monthly_reports:
            try:
                wb = openpyxl.load_workbook(report_path, data_only=True)
                ws = wb.active

                header_map = self._get_header_map(ws)
                name_col = header_map.get("nombre del perfil")
                executions_col = header_map.get("cantidad de ejecuciones")
                optimal_col = header_map.get("cantidad de ejecuciones recomendadas")
                automatic_col = header_map.get("bot automático")
                manual_col = header_map.get("bot manual")
                offline_col = header_map.get("bot offline")
                last_search_col = header_map.get("última búsqueda")
                responsable_col = header_map.get("responsable")
                last_update_col = header_map.get("última actualización")
                delivery_col = header_map.get("fecha de entrega")

                if not name_col or not executions_col or not optimal_col:
                    continue

                for row in range(5, ws.max_row + 1):
                    profile_name = ws.cell(row=row, column=name_col).value
                    if not profile_name:
                        continue

                    responsable = ws.cell(row=row, column=responsable_col).value if responsable_col else ""
                    executions = ws.cell(row=row, column=executions_col).value or 0
                    optimal_cell = ws.cell(row=row, column=optimal_col).value
                    is_automatic = bool(automatic_col and ws.cell(row=row, column=automatic_col).value == "X")
                    is_manual = bool(manual_col and ws.cell(row=row, column=manual_col).value == "X")
                    is_offline = bool(offline_col and ws.cell(row=row, column=offline_col).value == "X")
                    last_search = ws.cell(row=row, column=last_search_col).value if last_search_col else None
                    last_update_text = ws.cell(row=row, column=last_update_col).value if last_update_col else ""
                    delivery_date_text = ws.cell(row=row, column=delivery_col).value if delivery_col else ""

                    if isinstance(last_update_text, str) and last_update_text.strip() == "—":
                        last_update_text = ""
                    if isinstance(delivery_date_text, str) and delivery_date_text.strip() == "—":
                        delivery_date_text = ""

                    # Procesar ejecuciones óptimas
                    daily_optimal = self._extract_optimal_value(optimal_cell)
                    has_tracking = daily_optimal > 0

                    if profile_name not in aggregated_data:
                        aggregated_data[profile_name] = {
                            'executions': 0,
                            'daily_optimal': daily_optimal,
                            'monthly_optimal': daily_optimal * days_in_month if has_tracking else 0,
                            'has_tracking': has_tracking,
                            'is_automatic': is_automatic,
                            'is_manual': is_manual,
                            'is_offline': is_offline,
                            'last_search': last_search,
                            'responsable': responsable,
                            'last_update_text': last_update_text or "",
                            'delivery_date_text': delivery_date_text or ""
                        }
                    else:
                        if responsable and not aggregated_data[profile_name].get('responsable'):
                            aggregated_data[profile_name]['responsable'] = responsable
                        aggregated_data[profile_name]['is_automatic'] = (
                            aggregated_data[profile_name]['is_automatic'] or is_automatic
                        )
                        aggregated_data[profile_name]['is_manual'] = (
                            aggregated_data[profile_name]['is_manual'] or is_manual
                        )
                        aggregated_data[profile_name]['is_offline'] = (
                            aggregated_data[profile_name].get('is_offline', False) or is_offline
                        )

                        if last_update_text:
                            aggregated_data[profile_name]['last_update_text'] = last_update_text

                        if delivery_date_text:
                            aggregated_data[profile_name]['delivery_date_text'] = delivery_date_text

                    # Acumular ejecuciones
                    aggregated_data[profile_name]['executions'] += executions

                    # Actualizar última búsqueda si es más reciente
                    if last_search and (not aggregated_data[profile_name]['last_search'] or
                                        last_search > aggregated_data[profile_name]['last_search']):
                        aggregated_data[profile_name]['last_search'] = last_search

            except Exception as e:
                print(f"Error procesando archivo {report_path}: {e}")
                continue

        return {
            'aggregated_data': aggregated_data,
            'reports_found': len(monthly_reports),
            'reports_count': len(monthly_reports),
            'days_in_month': days_in_month
        }

    def _extract_optimal_value(self, optimal_cell):
        """Extrae el valor numérico de ejecuciones óptimas de la celda."""
        if not optimal_cell:
            return 0

        if isinstance(optimal_cell, (int, float)):
            return max(0, int(optimal_cell))

        if isinstance(optimal_cell, str):
            # Extraer número de strings como "🎯 30" o "◼ Deshabilitado"
            import re
            if "deshabilitado" in optimal_cell.lower() or "n/a" in optimal_cell.lower():
                return 0

            match = re.search(r'\d+', optimal_cell)
            if match:
                return int(match.group())

        return 0

    def _calculate_weekly_success_percentage(self, executions, weekly_optimal):
        """Calcula el porcentaje de éxito semanal correctamente."""
        if weekly_optimal <= 0:
            return None
        return (executions / weekly_optimal) * 100

    def _calculate_monthly_success_percentage(self, executions, monthly_optimal):
        """Calcula el porcentaje de éxito mensual correctamente."""
        if monthly_optimal <= 0:
            return None
        return (executions / monthly_optimal) * 100

    def _get_success_format(self, percentage):
        """Obtiene el formato apropiado basado en el porcentaje de éxito."""
        if percentage is None:
            return "N/A", None, None

        if percentage >= 100.0:
            return f"✅ {percentage:.1f}%", "FF90EE90", "006400"
        elif percentage >= 90.0:
            return f"📊 {percentage:.1f}%", "FFE6E6FA", "800080"
        elif percentage >= 50.0:
            return f"📊 {percentage:.1f}%", "FFE6E6FA", "800080"
        elif percentage >= 30.0:
            return f"⚠️ {percentage:.1f}%", "FFFFFF99", "B8860B"
        else:
            return f"❌ {percentage:.1f}%", "FFFFCCCC", "CC0000"

    def _add_weekly_profile_data(self, worksheet, aggregated_data, styles):
        """Agrega los datos de perfiles al reporte semanal con cálculos corregidos."""
        row_num = 5

        for profile_name, data in aggregated_data.items():
            # Nombre del Perfil
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = profile_name
            cell.border = styles['border']

            # Responsable
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = data.get('responsable') or "—"
            cell.border = styles['border']

            # Última Actualización
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = data.get('last_update_text') or "—"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Fecha de entrega
            cell = worksheet.cell(row=row_num, column=4)
            cell.value = data.get('delivery_date_text') or "—"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Ejecuciones Acumuladas
            cell = worksheet.cell(row=row_num, column=5)
            cell.value = data['executions']
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Ejecuciones Óptimas Semanales
            cell = worksheet.cell(row=row_num, column=6)
            if data['has_tracking']:
                cell.value = f"🎯 {data['weekly_optimal']} (7 días)"
            else:
                cell.value = "◼ Deshabilitado"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Porcentaje de Éxito Semanal (CORREGIDO)
            cell = worksheet.cell(row=row_num, column=7)
            success_percentage = self._calculate_weekly_success_percentage(
                data['executions'], data['weekly_optimal']
            )

            success_display, fill_color, font_color = self._get_success_format(success_percentage)
            cell.value = success_display
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            if fill_color and font_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(bold=True, color=font_color)

            # Bot Automático
            cell = worksheet.cell(row=row_num, column=8)
            if data['is_automatic']:
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Bot Manual
            cell = worksheet.cell(row=row_num, column=9)
            if data['is_manual']:
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Bot Offline
            cell = worksheet.cell(row=row_num, column=10)
            if data.get('is_offline'):
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Última Búsqueda
            cell = worksheet.cell(row=row_num, column=11)
            cell.value = data['last_search'] if data['last_search'] else "Nunca"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            row_num += 1

    def _add_monthly_profile_data(self, worksheet, aggregated_data, styles):
        """Agrega los datos de perfiles al reporte mensual con cálculos corregidos."""
        row_num = 5

        for profile_name, data in aggregated_data.items():
            # Nombre del Perfil
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = profile_name
            cell.border = styles['border']

            # Responsable
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = data.get('responsable') or "—"
            cell.border = styles['border']

            # Última Actualización
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = data.get('last_update_text') or "—"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Fecha de entrega
            cell = worksheet.cell(row=row_num, column=4)
            cell.value = data.get('delivery_date_text') or "—"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Ejecuciones Acumuladas
            cell = worksheet.cell(row=row_num, column=5)
            cell.value = data['executions']
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Ejecuciones Óptimas Mensuales
            cell = worksheet.cell(row=row_num, column=6)
            if data['has_tracking']:
                days_count = data['monthly_optimal'] // data['daily_optimal'] if data['daily_optimal'] > 0 else 0
                cell.value = f"🎯 {data['monthly_optimal']} ({days_count} días)"
            else:
                cell.value = "◼ Deshabilitado"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Porcentaje de Éxito Mensual
            cell = worksheet.cell(row=row_num, column=7)
            success_percentage = self._calculate_monthly_success_percentage(
                data['executions'], data['monthly_optimal']
            )

            success_display, fill_color, font_color = self._get_success_format(success_percentage)
            cell.value = success_display
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            if fill_color and font_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(bold=True, color=font_color)

            # Bot Automático
            cell = worksheet.cell(row=row_num, column=8)
            if data['is_automatic']:
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Bot Manual
            cell = worksheet.cell(row=row_num, column=9)
            if data['is_manual']:
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Bot Offline
            cell = worksheet.cell(row=row_num, column=10)
            if data.get('is_offline'):
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Última Búsqueda
            cell = worksheet.cell(row=row_num, column=11)
            cell.value = data['last_search'] if data['last_search'] else "Nunca"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            row_num += 1

    def _get_report_styles(self):
        """Define los estilos reutilizables para los reportes."""
        return {
            'title_font': Font(bold=True, size=16, color="FFFFFF"),
            'title_fill': PatternFill(start_color="FF2E5090", end_color="FF2E5090", fill_type="solid"),
            'title_alignment': Alignment(horizontal="center", vertical="center"),
            'subtitle_font': Font(bold=True, size=12, color="000000"),
            'subtitle_alignment': Alignment(horizontal="center", vertical="center"),
            'header_font': Font(bold=True, color="FFFFFF"),
            'header_fill': PatternFill(start_color="FF366092", end_color="FF366092", fill_type="solid"),
            'header_alignment': Alignment(horizontal="center", vertical="center"),
            'bot_fill': PatternFill(start_color="FFFFC0CB", end_color="FFFFC0CB", fill_type="solid"),
            'bot_font': Font(bold=True, color="C71585"),
            'border': Border(
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000")
            )
        }

    def _add_daily_header(self, worksheet, total_bots, styles):
        """Agrega el encabezado para reportes diarios."""
        worksheet.merge_cells('A1:K1')
        title_cell = worksheet['A1']
        title_cell.value = "Reporte de Ejecuciones - Registro Diario"
        title_cell.font = styles['title_font']
        title_cell.fill = styles['title_fill']
        title_cell.alignment = styles['title_alignment']
        title_cell.border = styles['border']

        for col in range(1, 12):
            worksheet.cell(row=1, column=col).border = styles['border']

        worksheet.merge_cells('A2:K2')
        subtitle_cell = worksheet['A2']
        current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        subtitle_cell.value = f"Generado el {current_date} - Total de Bots: {total_bots}"
        subtitle_cell.font = styles['subtitle_font']
        subtitle_cell.alignment = styles['subtitle_alignment']

        for col in range(1, 12):
            worksheet.cell(row=2, column=col).border = styles['border']

        worksheet.row_dimensions[3].height = 10

    def _add_weekly_header(self, worksheet, start_date, end_date, total_bots, styles):
        """Agrega el encabezado para reportes semanales."""
        week_number = start_date.isocalendar()[1]

        worksheet.merge_cells('A1:K1')
        title_cell = worksheet['A1']
        title_cell.value = f"Reporte de Ejecuciones - Resumen Semanal (Semana {week_number})"
        title_cell.font = styles['title_font']
        title_cell.fill = styles['title_fill']
        title_cell.alignment = styles['title_alignment']
        title_cell.border = styles['border']

        for col in range(1, 12):
            worksheet.cell(row=1, column=col).border = styles['border']

        worksheet.merge_cells('A2:K2')
        subtitle_cell = worksheet['A2']
        current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        period_text = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        subtitle_cell.value = f"Generado el {current_date} - Período: {period_text}"
        subtitle_cell.font = styles['subtitle_font']
        subtitle_cell.alignment = styles['subtitle_alignment']

        for col in range(1, 12):
            worksheet.cell(row=2, column=col).border = styles['border']

        worksheet.row_dimensions[3].height = 10

    def _add_monthly_header(self, worksheet, start_date, end_date, total_bots, styles):
        """Agrega el encabezado para reportes mensuales."""
        month_name = start_date.strftime("%B").capitalize()
        year = start_date.year

        worksheet.merge_cells('A1:K1')
        title_cell = worksheet['A1']
        title_cell.value = f"Reporte de Ejecuciones - Resumen Mensual ({month_name} {year})"
        title_cell.font = styles['title_font']
        title_cell.fill = styles['title_fill']
        title_cell.alignment = styles['title_alignment']
        title_cell.border = styles['border']

        for col in range(1, 12):
            worksheet.cell(row=1, column=col).border = styles['border']

        worksheet.merge_cells('A2:K2')
        subtitle_cell = worksheet['A2']
        current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        period_text = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        subtitle_cell.value = f"Generado el {current_date} - Período: {period_text}"
        subtitle_cell.font = styles['subtitle_font']
        subtitle_cell.alignment = styles['subtitle_alignment']

        for col in range(1, 12):
            worksheet.cell(row=2, column=col).border = styles['border']

        worksheet.row_dimensions[3].height = 10

    def _add_table_headers(self, worksheet, styles):
        """Agrega los encabezados de la tabla."""
        headers = [
            "Nombre del Perfil",
            "Responsable",
            "Última Actualización",
            "Fecha de entrega",
            "Cantidad de ejecuciones",
            "Cantidad de Ejecuciones recomendadas",
            "Porcentaje de Éxito",
            "Bot Automático",
            "Bot Manual",
            "Bot Offline",
            "Última Búsqueda"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_alignment']
            cell.border = styles['border']

    def _add_profile_data(self, worksheet, profiles, styles):
        """Agrega los datos de perfiles para reporte diario."""
        for row_num, profile in enumerate(profiles, 5):
            # Nombre del Perfil
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = profile.name
            cell.border = styles['border']

            # Responsable
            cell = worksheet.cell(row=row_num, column=2)
            cell.value = profile.responsable if getattr(profile, "responsable", "") else "—"
            cell.border = styles['border']

            # Última Actualización (texto manual)
            cell = worksheet.cell(row=row_num, column=3)
            cell.value = profile.get_last_update_display() if hasattr(profile, "get_last_update_display") else "—"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Fecha de entrega (texto manual)
            cell = worksheet.cell(row=row_num, column=4)
            cell.value = profile.get_delivery_date_display() if hasattr(profile, "get_delivery_date_display") else "—"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Cantidad de ejecuciones
            cell = worksheet.cell(row=row_num, column=5)
            cell.value = profile.found_emails
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Ejecuciones Óptimas
            cell = worksheet.cell(row=row_num, column=6)
            cell.value = profile.get_optimal_display()
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Porcentaje de Éxito
            cell = worksheet.cell(row=row_num, column=7)
            success_display = profile.get_success_display()
            success_percentage = profile.get_success_percentage()

            cell.value = success_display
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Aplicar formato condicional
            if success_percentage is not None:
                _, fill_color, font_color = self._get_success_format(success_percentage)
                if fill_color and font_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(bold=True, color=font_color)

            # Bot Automático
            cell = worksheet.cell(row=row_num, column=8)
            if profile.is_bot_automatic():
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Bot Manual
            cell = worksheet.cell(row=row_num, column=9)
            if profile.is_bot_manual():
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Bot Offline
            cell = worksheet.cell(row=row_num, column=10)
            if hasattr(profile, "is_bot_offline") and profile.is_bot_offline():
                cell.value = "X"
                cell.fill = styles['bot_fill']
                cell.font = styles['bot_font']
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

            # Última Búsqueda
            cell = worksheet.cell(row=row_num, column=11)
            if profile.last_search:
                cell.value = profile.last_search.strftime("%d/%m/%Y %H:%M:%S")
            else:
                cell.value = "Nunca"
            cell.alignment = Alignment(horizontal="center")
            cell.border = styles['border']

    def _format_daily_worksheet(self, worksheet):
        """Aplica formato general a la hoja de trabajo."""
        column_widths = {
            1: 30,
            2: 22,
            3: 24,
            4: 22,
            5: 20,
            6: 35,
            7: 18,
            8: 15,
            9: 12,
            10: 12,
            11: 22
        }

        for col_num, width in column_widths.items():
            worksheet.column_dimensions[get_column_letter(col_num)].width = width

        worksheet.row_dimensions[1].height = 25
        worksheet.row_dimensions[2].height = 20

    def _create_weekly_file(self, start_date, end_date):
        """Crea el nombre y ruta del archivo de reporte semanal."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        week_number = start_date.isocalendar()[1]
        filename = f"reporte_semanal_{timestamp}_semana{week_number}.xlsx"
        return self.reports_dir / filename

    def _create_monthly_file(self, start_date, end_date):
        """Crea el nombre y ruta del archivo de reporte mensual."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        month_name = start_date.strftime("%B").lower()
        year = start_date.year
        filename = f"reporte_mensual_{timestamp}_{month_name}{year}.xlsx"
        return self.reports_dir / filename

    def _add_summary_sheet(self, worksheet, profiles):
        """Agrega hoja de resumen para reporte diario."""
        worksheet.cell(row=1, column=1).value = "RESUMEN EJECUTIVO - DIARIO"
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=16, color="366092")
        worksheet.merge_cells('A1:B1')

        current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        worksheet.cell(row=3, column=1).value = "Fecha de generación:"
        worksheet.cell(row=3, column=1).font = Font(bold=True)
        worksheet.cell(row=3, column=2).value = current_date

        worksheet.cell(row=4, column=1).value = "Total de bots:"
        worksheet.cell(row=4, column=1).font = Font(bold=True)
        worksheet.cell(row=4, column=2).value = len(profiles)

        # Agregar más métricas del resumen...
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 30

    def _add_weekly_summary_sheet(self, worksheet, weekly_data, start_date, end_date):
        """Agrega hoja de resumen semanal con métricas corregidas."""
        worksheet.cell(row=1, column=1).value = "RESUMEN EJECUTIVO - SEMANAL"
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=16, color="366092")
        worksheet.merge_cells('A1:B1')

        current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        worksheet.cell(row=3, column=1).value = "Fecha de generación:"
        worksheet.cell(row=3, column=1).font = Font(bold=True)
        worksheet.cell(row=3, column=2).value = current_date

        period_text = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        worksheet.cell(row=4, column=1).value = "Período del reporte:"
        worksheet.cell(row=4, column=1).font = Font(bold=True)
        worksheet.cell(row=4, column=2).value = period_text

        worksheet.cell(row=5, column=1).value = "Reportes diarios incluidos:"
        worksheet.cell(row=5, column=1).font = Font(bold=True)
        worksheet.cell(row=5, column=2).value = weekly_data['reports_count']

        # Agregar métricas de éxito semanal corregidas
        profiles_with_tracking = [
            data for data in weekly_data['aggregated_data'].values()
            if data['has_tracking']
        ]

        worksheet.cell(row=7, column=1).value = "MÉTRICAS DE ÉXITO SEMANAL (CORREGIDAS)"
        worksheet.cell(row=7, column=1).font = Font(bold=True, size=14, color="006400")

        worksheet.cell(row=8, column=1).value = "Perfiles con seguimiento:"
        worksheet.cell(row=8, column=1).font = Font(bold=True)
        worksheet.cell(row=8, column=2).value = len(profiles_with_tracking)

        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 30

    def _add_monthly_summary_sheet(self, worksheet, monthly_data, start_date, end_date):
        """Agrega hoja de resumen mensual con métricas corregidas."""
        worksheet.cell(row=1, column=1).value = "RESUMEN EJECUTIVO - MENSUAL"
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=16, color="366092")
        worksheet.merge_cells('A1:B1')

        current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        worksheet.cell(row=3, column=1).value = "Fecha de generación:"
        worksheet.cell(row=3, column=1).font = Font(bold=True)
        worksheet.cell(row=3, column=2).value = current_date

        month_name = start_date.strftime("%B").capitalize()
        year = start_date.year
        period_text = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

        worksheet.cell(row=4, column=1).value = "Mes:"
        worksheet.cell(row=4, column=1).font = Font(bold=True)
        worksheet.cell(row=4, column=2).value = f"{month_name} {year}"

        worksheet.cell(row=5, column=1).value = "Período completo:"
        worksheet.cell(row=5, column=1).font = Font(bold=True)
        worksheet.cell(row=5, column=2).value = period_text

        worksheet.cell(row=6, column=1).value = "Días en el mes:"
        worksheet.cell(row=6, column=1).font = Font(bold=True)
        worksheet.cell(row=6, column=2).value = monthly_data['days_in_month']

        worksheet.cell(row=7, column=1).value = "Reportes diarios incluidos:"
        worksheet.cell(row=7, column=1).font = Font(bold=True)
        worksheet.cell(row=7, column=2).value = monthly_data['reports_count']

        # Agregar métricas de éxito mensual
        profiles_with_tracking = [
            data for data in monthly_data['aggregated_data'].values()
            if data['has_tracking']
        ]

        # Calcular métricas acumuladas mensuales
        total_executions = sum(data['executions'] for data in monthly_data['aggregated_data'].values())
        optimal_profiles = sum(1 for data in profiles_with_tracking
                               if self._calculate_monthly_success_percentage(data['executions'],
                                                                             data['monthly_optimal']) >= 100)

        worksheet.cell(row=9, column=1).value = "MÉTRICAS DE ÉXITO MENSUAL"
        worksheet.cell(row=9, column=1).font = Font(bold=True, size=14, color="006400")

        worksheet.cell(row=10, column=1).value = "Perfiles con seguimiento:"
        worksheet.cell(row=10, column=1).font = Font(bold=True)
        worksheet.cell(row=10, column=2).value = len(profiles_with_tracking)

        worksheet.cell(row=11, column=1).value = "Perfiles que alcanzaron objetivo mensual:"
        worksheet.cell(row=11, column=1).font = Font(bold=True)
        worksheet.cell(row=11, column=2).value = optimal_profiles

        worksheet.cell(row=12, column=1).value = "Total de ejecuciones en el mes:"
        worksheet.cell(row=12, column=1).font = Font(bold=True)
        worksheet.cell(row=12, column=2).value = total_executions

        # Calcular tasa de éxito general
        if profiles_with_tracking:
            success_rate = (optimal_profiles / len(profiles_with_tracking)) * 100
            worksheet.cell(row=13, column=1).value = "Tasa de éxito general:"
            worksheet.cell(row=13, column=1).font = Font(bold=True)
            worksheet.cell(row=13, column=2).value = f"{success_rate:.1f}%"

        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 30

    def get_reports_directory(self):
        """Retorna el directorio donde se guardan los reportes."""
        return str(self.reports_dir)